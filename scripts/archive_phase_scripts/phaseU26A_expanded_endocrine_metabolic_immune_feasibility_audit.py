#!/usr/bin/env python3
"""
Phase U26A - Expanded endocrine-metabolic-immune submodule feasibility audit.

This script is intentionally non-destructive. It does not modify manuscript or
figure files. It builds a curated submodule library, discovers dataset gene
universes, audits coverage, inspects metadata for feasible contrasts, and
writes figure-priority and research-question recommendations.
"""

from __future__ import annotations

import argparse
import csv
import gzip
import json
import math
import os
import re
import shutil
import statistics
import subprocess
import sys
import tempfile
from collections import Counter, defaultdict
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Sequence, Set, Tuple

SCRIPT_VERSION = "U26A_v1.0_2026-07-14"

DATASETS = {
    "GSE186800": {
        "label": "recurrence-trigger bladder model",
        "layer": "bulk",
        "accession": "GSE186800",
        "aliases": ["recurrence", "recurrent", "trigger", "bladder"],
        "expected_contrasts": [
            "recurrence-trigger or rechallenge state versus comparator",
            "infected versus uninfected/control",
        ],
    },
    "GSE280297": {
        "label": "pregnancy-associated UTI model",
        "layer": "bulk",
        "accession": "GSE280297",
        "aliases": ["pregnancy", "pregnant", "gestation", "maternal"],
        "expected_contrasts": [
            "pregnancy-associated UTI versus pregnancy comparator",
            "adverse pregnancy outcome versus non-adverse outcome, only if metadata supports it",
        ],
    },
    "GSE112098": {
        "label": "human urine inflammatory comparator",
        "layer": "bulk",
        "accession": "GSE112098",
        "aliases": ["urine", "urinary", "inflammatory", "comparator"],
        "expected_contrasts": [
            "UTI or inflammatory urine state versus comparator",
        ],
    },
    "GSE252321": {
        "label": "UPEC-responsive single-cell validation",
        "layer": "single-cell",
        "accession": "GSE252321",
        "aliases": ["upec", "single_cell", "single-cell", "scrna"],
        "expected_contrasts": [
            "UPEC versus control at sample level",
            "UPEC versus control within resolved cell populations",
        ],
    },
}

# Each tuple: axis, submodule_id, display_label, genes, biological_readout,
# priority_weight, figure_family.
SUBMODULES: List[Tuple[str, str, str, str, str, int, str]] = [
    ("steroid_cholesterol_endocrine", "CHOLESTEROL_BIOSYNTHESIS", "Cholesterol biosynthesis",
     "HMGCS1 HMGCR MVK PMVK MVD IDI1 FDPS FDFT1 SQLE LSS CYP51A1 TM7SF2 MSMO1 NSDHL HSD17B7 EBP SC5D DHCR7 DHCR24",
     "de novo cholesterol precursor production", 5, "Figure_7"),
    ("steroid_cholesterol_endocrine", "CHOLESTEROL_UPTAKE_TRANSPORT_EXPORT", "Cholesterol uptake, transport and export",
     "LDLR LRP1 SCARB1 CD36 NPC1 NPC2 STARD3 STARD4 STARD5 APOE APOA1 APOB ABCA1 ABCG1 ABCG5 ABCG8 OSBPL1A OSBPL2 OSBPL5",
     "cellular sterol acquisition, trafficking and efflux", 5, "Figure_7"),
    ("steroid_cholesterol_endocrine", "STEROIDOGENESIS_CORE", "Core steroidogenesis",
     "STAR CYP11A1 HSD3B1 HSD3B2 CYP17A1 CYP21A2 CYP11B1 CYP11B2 POR FDX1 FDXR CYB5A",
     "conversion of cholesterol into steroid intermediates", 5, "Figure_7"),
    ("steroid_cholesterol_endocrine", "PROGESTERONE_BIOSYNTHESIS_RESPONSE", "Progesterone biosynthesis and response",
     "STAR CYP11A1 HSD3B1 HSD3B2 AKR1C1 AKR1C2 PGR PGRMC1 PGRMC2 PAQR5 PAQR7 PAQR8 NCOA1 NCOA2 NCOA3 FKBP5 HAND2",
     "progesterone-supportive synthesis and receptor activity", 5, "Figure_7"),
    ("steroid_cholesterol_endocrine", "ESTROGEN_BIOSYNTHESIS", "Estrogen biosynthesis and interconversion",
     "CYP19A1 HSD17B1 HSD17B2 HSD17B7 HSD17B8 HSD17B12 CYP1A1 CYP1B1 COMT",
     "aromatization and estrogen interconversion", 5, "Figure_7"),
    ("steroid_cholesterol_endocrine", "ESTROGEN_RECEPTOR_RESPONSE", "Estrogen receptor response",
     "ESR1 ESR2 GPER1 GREB1 TFF1 PGR NRIP1 NCOA1 NCOA2 NCOA3 FOXA1 KLF4 CCND1",
     "canonical and membrane-associated estrogen response", 5, "Figure_7"),
    ("steroid_cholesterol_endocrine", "ANDROGEN_TESTOSTERONE_BIOSYNTHESIS", "Androgen and testosterone biosynthesis",
     "STAR CYP11A1 HSD3B1 HSD3B2 CYP17A1 HSD17B3 AKR1C3 CYB5A POR",
     "androgen precursor and testosterone production", 5, "Figure_7"),
    ("steroid_cholesterol_endocrine", "TESTOSTERONE_CONVERSION_AROMATIZATION", "Testosterone conversion and aromatization",
     "CYP19A1 SRD5A1 SRD5A2 SRD5A3 HSD17B2 AKR1C2 AKR1C3 UGT2B7 UGT2B15 UGT2B17 SULT2A1",
     "conversion to estrogens, dihydrotestosterone or inactive conjugates", 5, "Figure_7"),
    ("steroid_cholesterol_endocrine", "ANDROGEN_RECEPTOR_SIGNALING", "Androgen receptor signaling",
     "AR NCOA1 NCOA2 NCOA3 NCOR1 NCOR2 FKBP5 KLK3 TMPRSS2 SGK1 FOXA1",
     "androgen-responsive nuclear receptor program", 4, "Figure_7"),
    ("steroid_cholesterol_endocrine", "GLUCOCORTICOID_RESPONSE", "Glucocorticoid response",
     "NR3C1 HSD11B1 HSD11B2 FKBP5 TSC22D3 DUSP1 SGK1 PER1 KLF9 ZBTB16 NFKBIA GILZ TAT",
     "glucocorticoid sensing and anti-inflammatory or stress adaptation", 5, "Figure_7"),
    ("steroid_cholesterol_endocrine", "MINERALOCORTICOID_RESPONSE", "Mineralocorticoid response",
     "NR3C2 HSD11B2 SGK1 SCNN1A SCNN1B SCNN1G NEDD4L ATP1A1 ATP1B1 FXYD2 KCNJ1",
     "mineralocorticoid-linked ion transport and epithelial response", 4, "Figure_7"),
    ("steroid_cholesterol_endocrine", "STEROID_SULFATION_DESULFATION", "Steroid sulfation and desulfation",
     "SULT1E1 SULT2A1 SULT2B1 STS PAPSS1 PAPSS2 SLC26A2 SLC35B2 ARSB GALNS",
     "reversible steroid conjugation and bioavailability", 5, "Figure_7"),
    ("steroid_cholesterol_endocrine", "STEROID_CATABOLISM_DEACTIVATION", "Steroid catabolism and deactivation",
     "CYP3A4 CYP3A5 CYP1A1 CYP1B1 UGT1A1 UGT1A3 UGT2B7 UGT2B15 UGT2B17 AKR1C1 AKR1C2 HSD17B2 COMT",
     "steroid clearance, oxidation, reduction and conjugation", 5, "Figure_7"),
    ("steroid_cholesterol_endocrine", "PLACENTAL_STEROID_METABOLISM", "Placental steroid metabolism",
     "STAR CYP11A1 HSD3B1 CYP19A1 HSD17B1 HSD17B2 STS SULT1E1 HSD11B2 CRH NR3C1 PGR GPER1 CYP21A2",
     "placental steroid production, conversion and fetal glucocorticoid protection", 5, "Figure_7"),
    ("steroid_cholesterol_endocrine", "LXR_FXR_STEROL_SIGNALING", "LXR and FXR sterol signaling",
     "NR1H3 NR1H2 NR1H4 RXRA RXRB ABCA1 ABCG1 APOE SREBF1 CYP7A1 SHP NR0B2",
     "sterol-sensing nuclear receptor activity", 4, "Figure_7"),

    ("lipid_metabolism", "FATTY_ACID_SYNTHESIS", "Fatty-acid synthesis",
     "ACLY ACACA ACACB FASN SCD ELOVL1 ELOVL5 ELOVL6 ELOVL7 ACSL3 ACSL4 SREBF1 MLXIPL ME1",
     "lipogenic carbon allocation and membrane precursor synthesis", 5, "Figure_7"),
    ("lipid_metabolism", "FATTY_ACID_BETA_OXIDATION", "Fatty-acid beta-oxidation",
     "CPT1A CPT1B CPT1C CPT2 ACADL ACADM ACADS HADHA HADHB ECHS1 ETFA ETFB ETFDH ACOX1 PPARA PPARGC1A",
     "mitochondrial and peroxisomal fatty-acid oxidation", 5, "Figure_7"),
    ("lipid_metabolism", "PHOSPHOLIPID_METABOLISM", "Phospholipid metabolism",
     "CHKA CHKB PCYT1A PCYT1B CEPT1 CHPT1 LPCAT1 LPCAT2 LPCAT3 LPCAT4 PLA2G4A PLA2G6 PLD1 PLD2 PISD PEMT",
     "membrane phospholipid synthesis and remodeling", 4, "Figure_7"),
    ("lipid_metabolism", "SPHINGOLIPID_CERAMIDE_METABOLISM", "Sphingolipid and ceramide metabolism",
     "SPTLC1 SPTLC2 SPTLC3 KDSR CERS1 CERS2 CERS3 CERS4 CERS5 CERS6 DEGS1 SGMS1 SGMS2 SMPD1 SMPD2 SMPD3 ASAH1 SPHK1 SPHK2 SGPL1",
     "ceramide, sphingosine and sphingosine-1-phosphate balance", 5, "Figure_7"),
    ("lipid_metabolism", "EICOSANOID_PROSTAGLANDIN_METABOLISM", "Eicosanoid and prostaglandin metabolism",
     "PLA2G4A PTGS1 PTGS2 PTGES PTGES2 PTGES3 HPGDS PTGDS ALOX5 ALOX5AP ALOX12 ALOX15 LTA4H LTC4S TBXAS1 CYP4F2 HPGD",
     "inflammatory lipid mediator generation and termination", 5, "Figure_7"),
    ("lipid_metabolism", "LIPID_DROPLET_DYNAMICS", "Lipid-droplet dynamics",
     "PLIN1 PLIN2 PLIN3 PLIN4 PLIN5 DGAT1 DGAT2 SOAT1 SOAT2 PNPLA2 ABHD5 CIDEC CIDEA FITM2 LIPE",
     "neutral-lipid storage, mobilization and stress buffering", 4, "Figure_7"),
    ("lipid_metabolism", "PPAR_SREBP_LXR_REGULATION", "PPAR, SREBP and LXR regulation",
     "PPARA PPARD PPARG RXRA RXRB SREBF1 SREBF2 SCAP INSIG1 INSIG2 NR1H3 NR1H2 ABCA1 ABCG1 FABP4 LPL",
     "lipid-sensing transcriptional control", 5, "Figure_7"),
    ("lipid_metabolism", "FERROPTOSIS_LIPID_PEROXIDATION", "Ferroptosis-linked lipid peroxidation",
     "ACSL4 LPCAT3 ALOX15 ALOX12 GPX4 SLC7A11 AIFM2 GCH1 DHODH TFRC NCOA4 HMOX1 SAT1 FTH1 FTL SLC40A1",
     "oxidized phospholipid accumulation and ferroptosis defense", 5, "Figure_7"),

    ("carbohydrate_inflammatory_carbon", "GLYCOLYSIS", "Glycolysis",
     "HK1 HK2 GPI PFKP PFKL PFKM ALDOA ALDOB ALDOC TPI1 GAPDH PGK1 PGAM1 ENO1 ENO2 PKM LDHA SLC2A1",
     "transcriptionally inferred glycolytic program", 5, "Figure_8"),
    ("carbohydrate_inflammatory_carbon", "GLUCONEOGENESIS", "Gluconeogenesis",
     "PCK1 PCK2 G6PC G6PC2 G6PC3 FBP1 FBP2 PC SLC37A4 PGM1 MLXIPL FOXO1",
     "glucose-producing and glucose-6-phosphate handling program", 4, "Figure_8"),
    ("carbohydrate_inflammatory_carbon", "GLYCOGEN_SYNTHESIS", "Glycogen synthesis",
     "GYS1 GYS2 GBE1 UGP2 PGM1 PGM2 PPP1R3A PPP1R3B PPP1R3C PPP1R3D",
     "glycogen storage program", 4, "Figure_8"),
    ("carbohydrate_inflammatory_carbon", "GLYCOGENOLYSIS", "Glycogenolysis",
     "PYGL PYGM PYGB AGL PGM1 PHKA1 PHKA2 PHKB PHKG1 PHKG2 CALM1 CALM2",
     "glycogen mobilization program", 4, "Figure_8"),
    ("carbohydrate_inflammatory_carbon", "PENTOSE_PHOSPHATE_PATHWAY", "Pentose-phosphate pathway",
     "G6PD PGLS PGD RPIA RPE TALDO1 TKT TKTL1 PRPS1 PRPS2",
     "NADPH and ribose-phosphate-generating program", 5, "Figure_8"),
    ("carbohydrate_inflammatory_carbon", "TCA_OXPHOS", "TCA cycle and oxidative phosphorylation",
     "CS ACO2 IDH3A IDH3B IDH3G OGDH DLST SUCLG1 SUCLG2 SDHA SDHB SDHC SDHD FH MDH2 NDUFS1 NDUFS2 NDUFV1 UQCRC1 UQCRC2 COX4I1 ATP5F1A ATP5F1B",
     "mitochondrial oxidative carbon metabolism", 5, "Figure_8"),
    ("carbohydrate_inflammatory_carbon", "LACTATE_HIF1A_INFLAMMATORY_GLYCOLYSIS", "Lactate and HIF1A inflammatory glycolysis",
     "HIF1A EPAS1 EGLN1 EGLN3 PDK1 PDK3 LDHA SLC16A1 SLC16A3 VEGFA BNIP3 BNIP3L HK2 PFKFB3 PFKFB4",
     "hypoxia-responsive inflammatory carbohydrate-metabolism program", 5, "Figure_8"),
    ("carbohydrate_inflammatory_carbon", "AMPK_SIGNALING", "AMPK signaling",
     "PRKAA1 PRKAA2 PRKAB1 PRKAB2 PRKAG1 PRKAG2 PRKAG3 STK11 CAB39 STRADA CAMKK2 TSC2 ULK1 ACACA ACACB",
     "energy-stress sensing and catabolic counter-regulation", 5, "Figure_8"),
    ("carbohydrate_inflammatory_carbon", "MTOR_SIGNALING", "mTOR signaling",
     "MTOR RPTOR RICTOR MLST8 AKT1 TSC1 TSC2 RHEB RRAGA RRAGB RRAGC RRAGD LAMTOR1 LAMTOR2 LAMTOR3 LAMTOR4 LAMTOR5 EIF4EBP1 RPS6KB1",
     "nutrient-sensing anabolic signaling", 5, "Figure_8"),
    ("carbohydrate_inflammatory_carbon", "GLUCOSE_TRANSPORT", "Glucose transport",
     "SLC2A1 SLC2A2 SLC2A3 SLC2A4 SLC2A5 SLC2A6 SLC2A8 SLC2A10 SLC2A13 SLC2A14 SLC5A1 SLC5A2",
     "cellular glucose and hexose uptake", 4, "Figure_8"),
    ("carbohydrate_inflammatory_carbon", "AGE_RAGE_SIGNALING", "AGE-RAGE signaling",
     "AGER HMGB1 S100A8 S100A9 S100A12 S100B DIAPH1 NFKB1 RELA MAPK1 MAPK3 CYBB NOX1 RAC1 IL6 TNF",
     "glycation-associated inflammatory signaling", 4, "Figure_8"),

    ("insulin_irs_signaling", "INSULIN_RECEPTOR_IRS", "Insulin receptor and IRS signaling",
     "INSR IRS1 IRS2 IRS4 SHC1 GRB2 SOS1 PIK3CA PIK3CB PIK3CD PIK3R1 PDPK1 AKT1 AKT2 FOXO1 GSK3B TBC1D4 SLC2A4",
     "proximal insulin receptor signaling and glucose-handling response", 5, "Figure_8"),
    ("insulin_irs_signaling", "PI3K_AKT_SIGNALING", "PI3K-AKT signaling",
     "PIK3CA PIK3CB PIK3CD PIK3CG PIK3R1 PIK3R2 PIK3R5 AKT1 AKT2 AKT3 PDPK1 PTEN PHLPP1 PHLPP2 GSK3B FOXO1 FOXO3",
     "growth, survival and immunometabolic signal integration", 5, "Figure_8"),
    ("insulin_irs_signaling", "MTORC1_INSULIN_OUTPUT", "mTORC1 insulin output",
     "INSR IRS1 IRS2 AKT1 TSC1 TSC2 RHEB MTOR RPTOR MLST8 RPS6KB1 EIF4EBP1 DEPTOR",
     "insulin-linked anabolic output", 5, "Figure_8"),
    ("insulin_irs_signaling", "AMPK_INSULIN_COUNTERREGULATION", "AMPK counter-regulation of insulin signaling",
     "PRKAA1 PRKAA2 STK11 TSC2 RPTOR ACACA IRS1 FOXO1 PPARGC1A SIRT1",
     "energy-stress opposition to insulin-mTOR activity", 4, "Figure_8"),

    ("adipokine_signaling", "LEPTIN_SIGNALING", "Leptin signaling",
     "LEP LEPR JAK2 STAT3 STAT5A STAT5B SOCS3 PTPN1 SH2B1 PIAS3 MAPK1 MAPK3",
     "leptin-JAK-STAT inflammatory and metabolic signaling", 5, "Figure_8"),
    ("adipokine_signaling", "ADIPONECTIN_SIGNALING", "Adiponectin signaling",
     "ADIPOQ ADIPOR1 ADIPOR2 APPL1 APPL2 PRKAA1 PRKAA2 PPARA PPARG PPARGC1A CAMKK2",
     "adiponectin-linked insulin sensitization and fatty-acid oxidation", 5, "Figure_8"),
    ("adipokine_signaling", "RESISTIN_INFLAMMATORY_SIGNALING", "Resistin-associated inflammatory signaling",
     "RETN CAP1 TLR4 NFKB1 RELA IL6 TNF CCL2 MAPK1 MAPK3 SOCS3",
     "resistin-associated innate inflammatory activation", 4, "Figure_8"),
    ("adipokine_signaling", "ADIPOKINE_INFLAMMATORY_AXIS", "Inflammatory adipokine axis",
     "LEP RETN NAMPT LCN2 RBP4 FABP4 SERPINE1 CCL2 IL6 TNF CXCL8 SAA1 SAA2",
     "adipose-linked inflammatory mediator program", 5, "Figure_8"),
    ("adipokine_signaling", "PPARG_ADIPOMETABOLIC_REGULATION", "PPARG adipometabolic regulation",
     "PPARG RXRA CEBPA CEBPB FABP4 LPL PLIN1 ADIPOQ SLC2A4 CD36 ACSL1 DGAT1",
     "PPARG-centered adipometabolic differentiation and lipid handling", 4, "Figure_8"),

    ("amino_acid_metabolism", "TRYPTOPHAN_KYNURENINE", "Tryptophan-kynurenine metabolism",
     "IDO1 IDO2 TDO2 AFMID KMO KYNU HAAO QPRT AADAT CCBL1 CCBL2 IL4I1 ACMSD",
     "immunoregulatory tryptophan depletion and kynurenine production", 5, "Figure_9"),
    ("amino_acid_metabolism", "ARGININE_NO_UREA", "Arginine, nitric oxide and urea coupling",
     "ARG1 ARG2 NOS1 NOS2 NOS3 ASS1 ASL ODC1 CPS1 OTC SLC7A1 SLC7A2 SLC7A5 DDAH1 DDAH2 GCH1",
     "arginine allocation between nitric oxide, polyamine and urea pathways", 5, "Figure_9"),
    ("amino_acid_metabolism", "GLUTAMINE_GLUTAMATE", "Glutamine and glutamate metabolism",
     "GLS GLS2 GLUL SLC1A5 SLC38A1 SLC38A2 SLC38A5 GOT1 GOT2 GLUD1 GLUD2 GPT2 ASNS",
     "glutamine uptake, anaplerosis and nitrogen transfer", 5, "Figure_9"),
    ("amino_acid_metabolism", "SERINE_GLYCINE_ONE_CARBON", "Serine, glycine and one-carbon metabolism",
     "PHGDH PSAT1 PSPH SHMT1 SHMT2 GLDC AMT MTHFD1 MTHFD2 MTHFD1L MTHFR DHFR TYMS",
     "one-carbon supply for redox balance, methylation and nucleotide synthesis", 5, "Figure_9"),
    ("amino_acid_metabolism", "METHIONINE_SAM_METHYLATION", "Methionine and SAM metabolism",
     "MAT1A MAT2A MAT2B AHCY GNMT BHMT MTR MTRR DNMT1 DNMT3A DNMT3B AHCYL1 CBS CTH",
     "methyl-donor generation and transsulfuration coupling", 5, "Figure_9"),
    ("amino_acid_metabolism", "BRANCHED_CHAIN_AMINO_ACIDS", "Branched-chain amino-acid metabolism",
     "BCAT1 BCAT2 BCKDHA BCKDHB DBT DLD PPM1K IVD MCCC1 MCCC2 ACAD8 HIBCH",
     "BCAA catabolism and immunometabolic carbon entry", 4, "Figure_9"),
    ("amino_acid_metabolism", "AROMATIC_AMINO_ACID_METABOLISM", "Aromatic amino-acid metabolism",
     "PAH TH DDC TAT HPD HGD MAOA MAOB COMT AANAT TPH1 TPH2 HAAO",
     "tyrosine, phenylalanine and tryptophan-derived metabolite handling", 4, "Figure_9"),
    ("amino_acid_metabolism", "AMINO_ACID_TRANSPORT", "Amino-acid transport",
     "SLC7A5 SLC7A8 SLC7A11 SLC1A5 SLC38A1 SLC38A2 SLC38A3 SLC38A5 SLC43A1 SLC43A2 SLC3A2 SLC6A14 SLC7A1 SLC7A2",
     "cellular amino-acid acquisition and exchange", 5, "Figure_9"),

    ("catecholamine_stress_adjacent", "CATECHOLAMINE_BIOSYNTHESIS_DEGRADATION", "Catecholamine biosynthesis and degradation",
     "TH DDC DBH PNMT GCH1 SPR QDPR MAOA MAOB COMT SLC6A2 SLC18A1 SLC18A2",
     "host catecholamine production, transport and clearance", 4, "Figure_9"),
    ("catecholamine_stress_adjacent", "ADRENERGIC_STRESS_SIGNALING", "Adrenergic stress signaling",
     "ADRA1A ADRA1B ADRA1D ADRA2A ADRA2B ADRA2C ADRB1 ADRB2 ADRB3 GNAS ADCY3 ADCY5 ADCY6 PRKACA CREB1",
     "alpha- and beta-adrenergic stress response", 4, "Figure_9"),
    ("catecholamine_stress_adjacent", "CATECHOLAMINE_IRON_REDOX_INTERFACE", "Catecholamine, iron and redox interface",
     "HMOX1 FTH1 FTL TFRC SLC40A1 CP HP LTF LCN2 SLC11A1 NCOA4 MAOA MAOB COMT CYBB",
     "host iron sequestration and redox context adjacent to catecholamine biology", 5, "Figure_9"),

    ("nucleotide_nad_nitrogen", "PURINE_DE_NOVO_SYNTHESIS", "De novo purine synthesis",
     "PPAT GART PFAS PAICS ADSL ATIC IMPDH1 IMPDH2 GMPS ADSS1 ADSS2 PRPS1 PRPS2",
     "purine nucleotide production for proliferation and stress responses", 5, "Figure_9"),
    ("nucleotide_nad_nitrogen", "PURINE_SALVAGE", "Purine salvage",
     "HPRT1 APRT PNP ADA ADK DCK ENTPD1 ENTPD2 NT5E AMPD1 AMPD2 AMPD3",
     "purine recycling and extracellular nucleotide handling", 4, "Figure_9"),
    ("nucleotide_nad_nitrogen", "PURINE_DEGRADATION_URATE", "Purine degradation and urate",
     "ADA PNP XDH AOX1 AMPD1 AMPD2 AMPD3 NT5C2 SLC22A12 ABCG2 SLC2A9",
     "purine catabolism, xanthine oxidation and urate transport", 5, "Figure_9"),
    ("nucleotide_nad_nitrogen", "PYRIMIDINE_SYNTHESIS", "Pyrimidine synthesis",
     "CAD DHODH UMPS CTPS1 CTPS2 TYMS RRM1 RRM2 RRM2B TK1 CMPK1 NME1",
     "pyrimidine nucleotide generation", 5, "Figure_9"),
    ("nucleotide_nad_nitrogen", "PYRIMIDINE_DEGRADATION", "Pyrimidine degradation",
     "DPYD DPYS UPB1 UPP1 UPP2 TYMP DCTD CDA NT5C NT5C2",
     "pyrimidine turnover and nucleoside degradation", 4, "Figure_9"),
    ("nucleotide_nad_nitrogen", "NAD_METABOLISM", "NAD metabolism",
     "NAMPT NAPRT NMNAT1 NMNAT2 NMNAT3 NADSYN1 QPRT KYNU PARP1 PARP2 SIRT1 SIRT2 SIRT3 CD38 BST1 ENPP1",
     "NAD synthesis, consumption and inflammatory signaling", 5, "Figure_9"),
    ("nucleotide_nad_nitrogen", "UREA_CYCLE", "Urea cycle",
     "CPS1 OTC ASS1 ASL ARG1 ARG2 SLC25A15 SLC25A13 NAGS OAT",
     "nitrogen disposal and arginine regeneration", 5, "Figure_9"),
    ("nucleotide_nad_nitrogen", "NITRIC_OXIDE_SYNTHESIS_REGULATION", "Nitric-oxide synthesis and regulation",
     "NOS1 NOS2 NOS3 GCH1 GCHFR ASS1 ASL ARG1 ARG2 DDAH1 DDAH2 PRMT1 SLC7A1 SLC7A2",
     "nitric-oxide production and substrate/cofactor control", 5, "Figure_9"),
    ("nucleotide_nad_nitrogen", "XANTHINE_OXIDASE_OXIDATIVE_PURINE_CATABOLISM", "Xanthine oxidase and oxidative purine catabolism",
     "XDH AOX1 PNP ADA SOD1 SOD2 CAT GPX1 CYBA CYBB NCF1 NCF2 HMOX1",
     "purine-linked reactive oxygen production and antioxidant response", 5, "Figure_9"),

    ("complement_architecture", "COMPLEMENT_CLASSICAL", "Classical complement pathway",
     "C1QA C1QB C1QC C1R C1S C2 C4A C4B C3 SERPING1",
     "antibody- and C1q-linked complement initiation", 5, "Figure_10"),
    ("complement_architecture", "COMPLEMENT_LECTIN", "Lectin complement pathway",
     "MBL2 MASP1 MASP2 MASP3 FCN1 FCN2 FCN3 COLEC10 COLEC11 C2 C4A C4B",
     "pattern-recognition lectin pathway initiation", 5, "Figure_10"),
    ("complement_architecture", "COMPLEMENT_ALTERNATIVE", "Alternative complement pathway",
     "C3 CFB CFD CFP CFH CFI CFHR1 CFHR2 CFHR3 CFHR4 CFHR5",
     "alternative pathway initiation and amplification", 5, "Figure_10"),
    ("complement_architecture", "COMPLEMENT_C3_CONVERTASE_AMPLIFICATION", "C3 convertase and amplification",
     "C2 C3 C4A C4B CFB CFD CFP CR1 CFH CFI",
     "formation and regulation of C3-cleaving complexes", 5, "Figure_10"),
    ("complement_architecture", "COMPLEMENT_TERMINAL_MAC", "Terminal complement and membrane-attack complex",
     "C5 C6 C7 C8A C8B C8G C9 CLU VTN CD59",
     "terminal complement assembly and membrane attack", 5, "Figure_10"),
    ("complement_architecture", "COMPLEMENT_REGULATORS", "Complement regulators",
     "CFH CFI CD46 CD55 CD59 CR1 C4BPA C4BPB SERPING1 CLU VTN CR2",
     "protection from excessive complement activation", 5, "Figure_10"),
    ("complement_architecture", "COMPLEMENT_COAGULATION_CROSSTALK", "Complement-coagulation crosstalk",
     "C3 C5 F3 F2 F5 F7 F10 F11 PLG SERPINE1 THBD PROC PROS1 VWF KNG1 KLKB1",
     "coupling of complement, coagulation, fibrinolysis and endothelial activation", 5, "Figure_10"),
    ("complement_architecture", "COMPLEMENT_OPSONOPHAGOCYTOSIS", "Complement-opsonophagocytosis",
     "C3 C4B C1QA C1QB C1QC CR1 CR2 ITGAM ITGB2 FCGR1A FCGR2A FCGR3A LILRB1 TYROBP",
     "complement deposition and phagocyte recognition", 5, "Figure_10"),
    ("complement_architecture", "COMPLEMENT_C3A_C5A_SIGNALING", "C3a and C5a inflammatory signaling",
     "C3 C5 C3AR1 C5AR1 C5AR2 GNAI2 NLRP3 IL1B CXCL8 TNF MAPK1 MAPK3",
     "anaphylatoxin-driven leukocyte recruitment and inflammatory amplification", 5, "Figure_10"),

    ("immune_context_anchors", "TLR4_LPS_SIGNALING_ANCHOR", "TLR4-LPS signaling anchor",
     "LBP CD14 TLR4 LY96 TIRAP MYD88 IRAK4 IRAK1 TRAF6 TICAM1 TICAM2 TBK1 IRF3 NFKB1 RELA",
     "UPEC/LPS-responsive innate sensing context", 5, "Integrated_network"),
    ("immune_context_anchors", "NFKB_MAPK_INFLAMMATION_ANCHOR", "NF-kB and MAPK inflammation anchor",
     "NFKB1 NFKB2 RELA RELB MAPK1 MAPK3 MAPK8 MAPK14 IKBKB IKBKG NFKBIA TNF IL1B IL6 CXCL8 CCL2",
     "central inflammatory transcriptional context", 5, "Integrated_network"),
    ("immune_context_anchors", "NEUTROPHIL_NETOSIS_ANCHOR", "Neutrophil and NETosis anchor",
     "S100A8 S100A9 FCGR3B CSF3R CXCR2 FPR1 MPO ELANE PRTN3 CTSG PADI4 CYBB NCF1 NCF2 LTF LCN2",
     "neutrophil recruitment, granule activity and NET-associated machinery", 5, "Integrated_network"),
    ("immune_context_anchors", "OXIDATIVE_STRESS_NRF2_ANCHOR", "Oxidative stress and NRF2 anchor",
     "NFE2L2 KEAP1 HMOX1 NQO1 GCLC GCLM TXNRD1 SOD1 SOD2 CAT GPX1 PRDX1 SRXN1",
     "antioxidant and oxidative-stress adaptation", 5, "Integrated_network"),
    ("immune_context_anchors", "UROTHELIAL_BARRIER_REPAIR_ANCHOR", "Urothelial barrier and repair anchor",
     "UPK1A UPK1B UPK2 UPK3A KRT8 KRT18 KRT19 CLDN4 CLDN7 OCLN TJP1 EPCAM KRT14 KRT5 EGFR AREG",
     "urothelial differentiation, junctional integrity and repair", 5, "Integrated_network"),
    ("immune_context_anchors", "PREGNANCY_INFLAMMATION_ANCHOR", "Pregnancy-associated inflammation anchor",
     "IL1B IL6 TNF CXCL8 CCL2 PTGS2 HIF1A NFKB1 RELA CRH HSD11B2 PGR ESR1 LEPR SERPINE1",
     "pregnancy-linked endocrine-inflammatory context", 5, "Integrated_network"),
]

RESEARCH_QUESTIONS = [
    {
        "id": "RQ1",
        "question": "Do pregnancy-associated UTI states shift from progesterone-supportive and steroid-conversion programs toward androgenic, glucocorticoid, inflammatory or steroid-catabolic programs?",
        "axes": ["steroid_cholesterol_endocrine", "immune_context_anchors"],
        "impact": 5,
        "figure": "Figure_7",
    },
    {
        "id": "RQ2",
        "question": "Does UTI induce an inflammatory carbon-use state integrating glycolysis, lactate-HIF1A, pentose-phosphate, insulin-IRS and mTOR signaling while suppressing oxidative metabolism?",
        "axes": ["carbohydrate_inflammatory_carbon", "insulin_irs_signaling"],
        "impact": 5,
        "figure": "Figure_8",
    },
    {
        "id": "RQ3",
        "question": "Which complement initiation and amplification route dominates recurrent UTI, pregnancy-associated UTI, urine inflammatory states and UPEC-responsive single cells?",
        "axes": ["complement_architecture", "immune_context_anchors"],
        "impact": 5,
        "figure": "Figure_10",
    },
    {
        "id": "RQ4",
        "question": "Do lipid remodeling, inflammatory eicosanoids and ferroptosis-linked lipid peroxidation connect UTI inflammation with altered steroid precursor availability?",
        "axes": ["lipid_metabolism", "steroid_cholesterol_endocrine"],
        "impact": 5,
        "figure": "Figure_7",
    },
    {
        "id": "RQ5",
        "question": "Do insulin, IRS and adipokine programs link pregnancy-associated inflammation, carbohydrate metabolism and recurrent UTI susceptibility?",
        "axes": ["insulin_irs_signaling", "adipokine_signaling", "carbohydrate_inflammatory_carbon"],
        "impact": 5,
        "figure": "Figure_8",
    },
    {
        "id": "RQ6",
        "question": "Do tryptophan-kynurenine, arginine-nitric-oxide, glutamine and one-carbon programs define immune activation and nutrient-handling states relevant to UPEC persistence?",
        "axes": ["amino_acid_metabolism", "nucleotide_nad_nitrogen"],
        "impact": 5,
        "figure": "Figure_9",
    },
    {
        "id": "RQ7",
        "question": "Are catecholamine-adjacent stress, iron-sequestration and redox programs linked to UPEC-responsive inflammatory biology?",
        "axes": ["catecholamine_stress_adjacent", "immune_context_anchors"],
        "impact": 4,
        "figure": "Figure_9_or_supplement",
    },
    {
        "id": "RQ8",
        "question": "Does UTI activate nucleotide turnover, NAD consumption, purine oxidation and nitrogen-handling programs that track inflammatory-cell proliferation and oxidative stress?",
        "axes": ["nucleotide_nad_nitrogen", "immune_context_anchors"],
        "impact": 4,
        "figure": "Figure_9",
    },
]

INDEX_BLUEPRINTS = [
    {
        "index_id": "STEROID_INFLAMMATORY_IMBALANCE_INDEX",
        "positive_modules": [
            "ANDROGEN_RECEPTOR_SIGNALING", "GLUCOCORTICOID_RESPONSE",
            "STEROID_CATABOLISM_DEACTIVATION", "NFKB_MAPK_INFLAMMATION_ANCHOR",
            "PREGNANCY_INFLAMMATION_ANCHOR",
        ],
        "negative_modules": [
            "PROGESTERONE_BIOSYNTHESIS_RESPONSE", "ESTROGEN_RECEPTOR_RESPONSE",
            "TESTOSTERONE_CONVERSION_AROMATIZATION", "PLACENTAL_STEROID_METABOLISM",
        ],
        "interpretation": "higher values indicate inflammatory/steroid-catabolic dominance relative to progesterone-estrogen supportive programs",
    },
    {
        "index_id": "INFLAMMATORY_CARBON_USE_INDEX",
        "positive_modules": [
            "GLYCOLYSIS", "LACTATE_HIF1A_INFLAMMATORY_GLYCOLYSIS",
            "PENTOSE_PHOSPHATE_PATHWAY", "MTOR_SIGNALING",
        ],
        "negative_modules": ["TCA_OXPHOS", "AMPK_SIGNALING"],
        "interpretation": "higher values indicate a transcriptionally inferred inflammatory carbohydrate-metabolism state",
    },
    {
        "index_id": "COMPLEMENT_INFLAMMATORY_AMPLIFICATION_INDEX",
        "positive_modules": [
            "COMPLEMENT_C3_CONVERTASE_AMPLIFICATION", "COMPLEMENT_TERMINAL_MAC",
            "COMPLEMENT_C3A_C5A_SIGNALING", "NFKB_MAPK_INFLAMMATION_ANCHOR",
            "NEUTROPHIL_NETOSIS_ANCHOR",
        ],
        "negative_modules": ["COMPLEMENT_REGULATORS"],
        "interpretation": "higher values indicate complement amplification and inflammatory effector dominance relative to regulation",
    },
    {
        "index_id": "PREGNANCY_RISK_ENDOCRINE_METABOLIC_INFLAMMATION_INDEX",
        "positive_modules": [
            "PREGNANCY_INFLAMMATION_ANCHOR", "ADIPOKINE_INFLAMMATORY_AXIS",
            "LACTATE_HIF1A_INFLAMMATORY_GLYCOLYSIS", "COMPLEMENT_C3A_C5A_SIGNALING",
            "STEROID_CATABOLISM_DEACTIVATION",
        ],
        "negative_modules": [
            "PROGESTERONE_BIOSYNTHESIS_RESPONSE", "PLACENTAL_STEROID_METABOLISM",
            "ADIPONECTIN_SIGNALING", "AMPK_SIGNALING",
        ],
        "interpretation": "higher values indicate endocrine-metabolic-inflammatory imbalance for outcome testing only where pregnancy metadata is adequate",
    },
]

ALLOWED_SUFFIXES = {
    ".tsv", ".csv", ".txt", ".gz", ".h5ad", ".h5", ".hdf5", ".rds",
    ".rda", ".rdata", ".xlsx", ".xls", ".parquet", ".feather", ".mtx",
}

IGNORE_PARTS = {
    ".git", "node_modules", "manuscript", "manuscripts", "figures", "figure",
    "plots", "plot", "archive", "archives", "submission", "submissions",
    "phaseu26a_expanded_endocrine_metabolic_immune_feasibility",
}

GENE_COL_KEYWORDS = (
    "gene_symbol", "genesymbol", "symbol", "gene_name", "genename", "gene",
    "feature_name", "feature", "external_gene_name", "hgnc_symbol",
)

META_KEYWORDS = (
    "meta", "metadata", "phenotype", "pheno", "clinical", "sample", "samples",
    "coldata", "obs", "annotation", "design", "group", "condition",
)

CONTRAST_KEYWORDS = (
    "group", "condition", "status", "case", "control", "infection", "infected",
    "upec", "uti", "preg", "gest", "trimester", "outcome", "miscar", "loss",
    "preterm", "recurr", "challenge", "treatment", "time", "cell_type", "celltype",
    "cluster", "phenotype", "disease",
)


def eprint(*args: object) -> None:
    print(*args, file=sys.stderr)


def normalize_gene_token(value: str) -> List[str]:
    if value is None:
        return []
    raw = str(value).strip().strip('"').strip("'")
    if not raw:
        return []
    raw = re.sub(r"\.[0-9]+$", "", raw)
    parts = re.split(r"[|;,/\s]+", raw)
    out: List[str] = []
    for token in parts:
        token = token.strip().upper()
        if not token or token in {"NA", "N/A", "NULL", "NONE", "GENE", "SYMBOL"}:
            continue
        if re.fullmatch(r"ENSG[0-9]{6,}", token):
            out.append(token)
        elif re.fullmatch(r"[A-Z][A-Z0-9-]{1,24}", token):
            out.append(token)
    return out


def is_ensembl(gene: str) -> bool:
    return bool(re.fullmatch(r"ENSG[0-9]{6,}", gene))


def open_text(path: Path):
    if path.suffix.lower() == ".gz":
        return gzip.open(path, "rt", encoding="utf-8", errors="replace")
    return path.open("r", encoding="utf-8", errors="replace")


def detect_delimiter(sample_lines: Sequence[str]) -> str:
    joined = "\n".join(sample_lines[:10])
    counts = {"\t": joined.count("\t"), ",": joined.count(","), ";": joined.count(";")}
    return max(counts, key=counts.get) if max(counts.values()) > 0 else "\t"


def infer_gene_column(header: List[str], path: Path) -> int:
    lower = [re.sub(r"[^a-z0-9]+", "_", x.strip().lower()).strip("_") for x in header]
    for key in GENE_COL_KEYWORDS:
        if key in lower:
            return lower.index(key)
    for i, col in enumerate(lower):
        if any(key in col for key in GENE_COL_KEYWORDS):
            return i
    if "feature" in path.name.lower() and len(header) >= 2:
        return 1
    return 0


def extract_genes_from_delimited(path: Path, max_lines: int = 300000) -> Set[str]:
    genes: Set[str] = set()
    try:
        with open_text(path) as handle:
            probe: List[str] = []
            for _ in range(20):
                line = handle.readline()
                if not line:
                    break
                if line.strip() and not line.startswith("#"):
                    probe.append(line.rstrip("\n\r"))
            if not probe:
                return genes
            delimiter = detect_delimiter(probe)
            rows = [next(csv.reader([line], delimiter=delimiter)) for line in probe]
            header = rows[0]
            gene_col = infer_gene_column(header, path)
            normalized_header = [x.strip().lower() for x in header]
            header_is_data = not any(any(k in x for k in GENE_COL_KEYWORDS) for x in normalized_header)
            data_rows = rows if header_is_data else rows[1:]
            for row in data_rows:
                if gene_col < len(row):
                    genes.update(normalize_gene_token(row[gene_col]))
            line_count = len(probe)
            for line in handle:
                line_count += 1
                if line_count > max_lines:
                    break
                if not line.strip() or line.startswith("#"):
                    continue
                row = next(csv.reader([line.rstrip("\n\r")], delimiter=delimiter))
                if gene_col < len(row):
                    genes.update(normalize_gene_token(row[gene_col]))
    except Exception:
        return set()
    return genes


def extract_genes_from_xlsx(path: Path, max_rows: int = 250000) -> Set[str]:
    try:
        from openpyxl import load_workbook  # type: ignore
    except Exception:
        return set()
    genes: Set[str] = set()
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
        for ws in wb.worksheets[:5]:
            iterator = ws.iter_rows(values_only=True)
            first = next(iterator, None)
            if first is None:
                continue
            header = ["" if x is None else str(x) for x in first]
            gene_col = infer_gene_column(header, path)
            for idx, row in enumerate(iterator, start=1):
                if idx > max_rows:
                    break
                if gene_col < len(row) and row[gene_col] is not None:
                    genes.update(normalize_gene_token(str(row[gene_col])))
        wb.close()
    except Exception:
        return set()
    return genes


def extract_genes_from_parquet(path: Path) -> Set[str]:
    try:
        import pandas as pd  # type: ignore
        df = pd.read_parquet(path)
        genes: Set[str] = set()
        if df.index is not None:
            for value in df.index.astype(str).tolist():
                genes.update(normalize_gene_token(value))
        for col in df.columns:
            if any(k in str(col).lower() for k in GENE_COL_KEYWORDS):
                for value in df[col].dropna().astype(str).tolist():
                    genes.update(normalize_gene_token(value))
        return genes
    except Exception:
        return set()


def extract_h5ad(path: Path) -> Tuple[Set[str], List[Dict[str, object]]]:
    genes: Set[str] = set()
    meta_rows: List[Dict[str, object]] = []
    try:
        import anndata as ad  # type: ignore
        obj = ad.read_h5ad(path, backed="r")
        genes = {g for value in obj.var_names.astype(str) for g in normalize_gene_token(value)}
        obs = obj.obs
        for col in obs.columns:
            vals = obs[col].astype(str).fillna("NA")
            counts = vals.value_counts(dropna=False)
            if 2 <= len(counts) <= 20:
                for value, count in counts.items():
                    meta_rows.append({"column": str(col), "value": str(value), "count": int(count)})
        if getattr(obj, "file", None) is not None:
            obj.file.close()
        return genes, meta_rows
    except Exception:
        pass
    try:
        import h5py  # type: ignore
        with h5py.File(path, "r") as h5:
            candidates = ["var/_index", "var_names", "raw/var/_index", "matrix/features/name", "matrix/features/id"]
            for key in candidates:
                if key in h5:
                    values = h5[key][()]
                    for value in values:
                        if isinstance(value, bytes):
                            value = value.decode("utf-8", errors="replace")
                        genes.update(normalize_gene_token(str(value)))
                    if genes:
                        break
    except Exception:
        pass
    return genes, meta_rows


def write_r_helper(path: Path) -> None:
    code = r'''
args <- commandArgs(trailingOnly=TRUE)
p <- args[1]
ext <- tolower(tools::file_ext(p))
obj <- NULL
if (ext == "rds") {
  obj <- tryCatch(readRDS(p), error=function(e) NULL)
} else {
  env <- new.env(parent=emptyenv())
  ok <- tryCatch({load(p, envir=env); TRUE}, error=function(e) FALSE)
  if (ok) {
    nms <- ls(env)
    if (length(nms) > 0) obj <- get(nms[1], envir=env)
  }
}
if (is.null(obj)) quit(status=0)
features <- NULL
meta <- NULL
if (inherits(obj, "Seurat")) {
  features <- tryCatch(rownames(obj), error=function(e) NULL)
  meta <- tryCatch(obj@meta.data, error=function(e) NULL)
} else if (inherits(obj, "SingleCellExperiment") || inherits(obj, "SummarizedExperiment")) {
  features <- tryCatch(rownames(obj), error=function(e) NULL)
  meta <- tryCatch(as.data.frame(SummarizedExperiment::colData(obj)), error=function(e) NULL)
} else if (is.matrix(obj) || is.data.frame(obj)) {
  features <- rownames(obj)
} else if (is.list(obj)) {
  preferred <- c("counts","count","expression","expr","matrix","data","normalized","vst")
  for (nm in preferred) {
    if (!is.null(obj[[nm]]) && (is.matrix(obj[[nm]]) || is.data.frame(obj[[nm]]))) {
      features <- rownames(obj[[nm]])
      break
    }
  }
  for (nm in c("metadata","meta","pheno","phenotype","coldata","sample_metadata")) {
    if (!is.null(obj[[nm]]) && is.data.frame(obj[[nm]])) {
      meta <- obj[[nm]]
      break
    }
  }
}
if (!is.null(features)) {
  features <- unique(as.character(features))
  features <- features[!is.na(features) & nzchar(features)]
  for (g in features) cat("G\t", g, "\n", sep="")
}
if (!is.null(meta) && nrow(meta) > 0) {
  for (nm in colnames(meta)) {
    v <- as.character(meta[[nm]])
    v[is.na(v)] <- "NA"
    tab <- sort(table(v), decreasing=TRUE)
    if (length(tab) >= 2 && length(tab) <= 20) {
      for (i in seq_along(tab)) cat("M\t", nm, "\t", names(tab)[i], "\t", as.integer(tab[i]), "\n", sep="")
    }
  }
}
'''
    path.write_text(code, encoding="utf-8")


def extract_r_object(path: Path, helper: Path) -> Tuple[Set[str], List[Dict[str, object]], str]:
    if shutil.which("Rscript") is None:
        return set(), [], "Rscript_not_available"
    genes: Set[str] = set()
    meta: List[Dict[str, object]] = []
    try:
        result = subprocess.run(
            ["Rscript", str(helper), str(path)],
            capture_output=True,
            text=True,
            timeout=240,
            check=False,
        )
        for line in result.stdout.splitlines():
            parts = line.split("\t")
            if not parts:
                continue
            if parts[0] == "G" and len(parts) >= 2:
                genes.update(normalize_gene_token(parts[1]))
            elif parts[0] == "M" and len(parts) >= 4:
                try:
                    count = int(parts[3])
                except ValueError:
                    count = 0
                meta.append({"column": parts[1], "value": parts[2], "count": count})
        status = "ok" if result.returncode == 0 else f"Rscript_exit_{result.returncode}"
        return genes, meta, status
    except subprocess.TimeoutExpired:
        return set(), [], "Rscript_timeout"
    except Exception as exc:
        return set(), [], f"Rscript_error:{type(exc).__name__}"


def parse_gtf_mappings(root: Path, max_files: int = 20) -> Dict[str, str]:
    mapping: Dict[str, str] = {}
    gtf_files = [p for p in root.rglob("*") if p.is_file() and (p.name.lower().endswith(".gtf") or p.name.lower().endswith(".gtf.gz"))]
    for path in gtf_files[:max_files]:
        try:
            with open_text(path) as handle:
                for line in handle:
                    if not line or line.startswith("#"):
                        continue
                    parts = line.rstrip("\n").split("\t")
                    if len(parts) < 9:
                        continue
                    attrs = parts[8]
                    gid = re.search(r'gene_id "([^"]+)"', attrs)
                    gname = re.search(r'gene_name "([^"]+)"', attrs)
                    if gid and gname:
                        ens = re.sub(r"\.[0-9]+$", "", gid.group(1)).upper()
                        symbols = normalize_gene_token(gname.group(1))
                        if ens.startswith("ENSG") and symbols:
                            mapping[ens] = symbols[0]
        except Exception:
            continue
    return mapping


def map_ensembl_with_orgdb(ensembl_ids: Set[str], output_dir: Path) -> Dict[str, str]:
    if not ensembl_ids or shutil.which("Rscript") is None:
        return {}
    input_file = output_dir / "_temporary_ensembl_ids.txt"
    output_file = output_dir / "_temporary_ensembl_map.tsv"
    r_file = output_dir / "_temporary_ensembl_map.R"
    input_file.write_text("\n".join(sorted(ensembl_ids)) + "\n", encoding="utf-8")
    r_code = r'''
args <- commandArgs(trailingOnly=TRUE)
infile <- args[1]
outfile <- args[2]
if (!requireNamespace("AnnotationDbi", quietly=TRUE) || !requireNamespace("org.Hs.eg.db", quietly=TRUE)) quit(status=0)
ids <- unique(readLines(infile, warn=FALSE))
ids <- sub("\\.[0-9]+$", "", ids)
res <- AnnotationDbi::mapIds(org.Hs.eg.db::org.Hs.eg.db, keys=ids, keytype="ENSEMBL", column="SYMBOL", multiVals="first")nout <- data.frame(ENSEMBL=names(res), SYMBOL=as.character(res), stringsAsFactors=FALSE)
out <- out[!is.na(out$SYMBOL) & nzchar(out$SYMBOL),]
write.table(out, outfile, sep="\t", row.names=FALSE, quote=FALSE)
'''
    # Repair a deliberate line split safely.
    r_code = r_code.replace(")nout <-", ")\nout <-")
    r_file.write_text(r_code, encoding="utf-8")
    mapping: Dict[str, str] = {}
    try:
        subprocess.run(["Rscript", str(r_file), str(input_file), str(output_file)], timeout=240, check=False)
        if output_file.exists():
            with output_file.open("r", encoding="utf-8", errors="replace") as handle:
                reader = csv.DictReader(handle, delimiter="\t")
                for row in reader:
                    ens = str(row.get("ENSEMBL", "")).upper()
                    sym = normalize_gene_token(str(row.get("SYMBOL", "")))
                    if ens and sym:
                        mapping[ens] = sym[0]
    except Exception:
        pass
    for p in (input_file, output_file, r_file):
        try:
            p.unlink()
        except OSError:
            pass
    return mapping


def file_is_ignored(path: Path) -> bool:
    lower_parts = {part.lower() for part in path.parts}
    if lower_parts.intersection(IGNORE_PARTS):
        return True
    if path.name.startswith("."):
        return True
    return False


def candidate_score(path: Path) -> int:
    name = path.name.lower()
    score = 0
    for key in ("gene", "feature", "expr", "expression", "count", "matrix", "normalized", "vst", "seurat", "sce", "h5ad"):
        if key in name:
            score += 3
    for key in META_KEYWORDS:
        if key in name:
            score += 1
    if path.suffix.lower() in {".h5ad", ".rds", ".rda", ".rdata"}:
        score += 4
    if path.suffix.lower() in {".png", ".pdf", ".svg", ".docx", ".pptx", ".zip"}:
        score -= 20
    return score


def discover_dataset_files(root: Path, dataset_id: str, spec: Dict[str, object]) -> List[Path]:
    all_files: List[Path] = []
    accession = str(spec["accession"]).lower()
    aliases = [str(x).lower() for x in spec["aliases"]]
    accession_hits: List[Path] = []
    alias_hits: List[Path] = []
    for path in root.rglob("*"):
        if not path.is_file() or file_is_ignored(path):
            continue
        name_lower = path.name.lower()
        suffix = path.suffix.lower()
        suffix2 = "".join(path.suffixes[-2:]).lower() if len(path.suffixes) >= 2 else suffix
        if suffix not in ALLOWED_SUFFIXES and suffix2 not in {".txt.gz", ".tsv.gz", ".csv.gz", ".mtx.gz"}:
            continue
        full = str(path).lower()
        if accession in full:
            accession_hits.append(path)
        else:
            alias_count = sum(1 for alias in aliases if alias in full)
            if alias_count >= 2:
                alias_hits.append(path)
    all_files = accession_hits if accession_hits else alias_hits
    return sorted(set(all_files), key=lambda p: (-candidate_score(p), len(str(p)), str(p)))


def extract_file(path: Path, r_helper: Path) -> Tuple[Set[str], List[Dict[str, object]], str]:
    lower = path.name.lower()
    suffix = path.suffix.lower()
    if suffix == ".h5ad":
        genes, meta = extract_h5ad(path)
        return genes, meta, "ok" if genes or meta else "unreadable_or_empty"
    if suffix in {".rds", ".rda", ".rdata"}:
        return extract_r_object(path, r_helper)
    if suffix in {".xlsx", ".xls"}:
        genes = extract_genes_from_xlsx(path)
        return genes, [], "ok" if genes else "unreadable_or_no_genes"
    if suffix in {".parquet", ".feather"}:
        genes = extract_genes_from_parquet(path)
        return genes, [], "ok" if genes else "unreadable_or_no_genes"
    if lower.endswith((".tsv", ".csv", ".txt", ".tsv.gz", ".csv.gz", ".txt.gz", ".features.tsv", ".features.tsv.gz")):
        genes = extract_genes_from_delimited(path)
        return genes, [], "ok" if genes else "unreadable_or_no_genes"
    if lower.endswith(("features.tsv", "features.tsv.gz", "genes.tsv", "genes.tsv.gz")):
        genes = extract_genes_from_delimited(path)
        return genes, [], "ok" if genes else "unreadable_or_no_genes"
    return set(), [], "unsupported_for_gene_extraction"


def read_metadata_table(path: Path, max_rows: int = 10000) -> List[Dict[str, object]]:
    rows: List[Dict[str, object]] = []
    try:
        with open_text(path) as handle:
            probe: List[str] = []
            for _ in range(10):
                line = handle.readline()
                if not line:
                    break
                if line.strip() and not line.startswith("#"):
                    probe.append(line.rstrip("\n\r"))
            if len(probe) < 2:
                return rows
            delimiter = detect_delimiter(probe)
            reader = csv.DictReader(probe + list(_take_lines(handle, max_rows)), delimiter=delimiter)
            counters: Dict[str, Counter] = defaultdict(Counter)
            for row in reader:
                for col, value in row.items():
                    if col is None:
                        continue
                    value = "NA" if value is None or str(value).strip() == "" else str(value).strip()
                    if len(value) <= 120:
                        counters[col][value] += 1
            for col, counts in counters.items():
                if 2 <= len(counts) <= 20:
                    for value, count in counts.most_common():
                        rows.append({"column": col, "value": value, "count": count})
    except Exception:
        return []
    return rows


def _take_lines(handle, max_rows: int) -> Iterable[str]:
    for idx, line in enumerate(handle):
        if idx >= max_rows:
            break
        yield line


def metadata_to_contrasts(dataset_id: str, meta_rows: List[Dict[str, object]], source: str) -> List[Dict[str, object]]:
    by_col: Dict[str, List[Tuple[str, int]]] = defaultdict(list)
    for row in meta_rows:
        try:
            count = int(row["count"])
        except Exception:
            count = 0
        by_col[str(row["column"])].append((str(row["value"]), count))
    out: List[Dict[str, object]] = []
    for col, values in by_col.items():
        col_lower = col.lower()
        if not any(key in col_lower for key in CONTRAST_KEYWORDS):
            continue
        values = sorted(values, key=lambda x: -x[1])
        n_levels = len(values)
        total_n = sum(x[1] for x in values)
        min_n = min((x[1] for x in values), default=0)
        if n_levels == 2 and total_n >= 30 and min_n >= 10:
            stats = "effect size, FDR, logistic regression/odds ratio, AUROC and permutation testing are potentially feasible"
            support = "strong_binary_metadata_candidate"
        elif n_levels == 2 and total_n >= 16 and min_n >= 5:
            stats = "effect size, FDR and AUROC/permutation are preferable; odds ratios may be unstable"
            support = "moderate_binary_metadata_candidate"
        elif 3 <= n_levels <= 10 and total_n >= 20:
            stats = "multi-group effect sizes, omnibus testing and pairwise FDR-controlled contrasts are potentially feasible"
            support = "multigroup_metadata_candidate"
        else:
            stats = "descriptive or exploratory analysis only unless sample-level metadata are improved"
            support = "limited_metadata_candidate"
        out.append({
            "dataset": dataset_id,
            "source": source,
            "metadata_column": col,
            "levels": " | ".join(f"{v}:{n}" for v, n in values),
            "n_levels": n_levels,
            "total_n": total_n,
            "minimum_level_n": min_n,
            "support_class": support,
            "recommended_statistics": stats,
        })
    return out


def submodule_records() -> List[Dict[str, object]]:
    records: List[Dict[str, object]] = []
    for axis, sid, label, genes, readout, weight, figure in SUBMODULES:
        gene_list = sorted(set(normalize_gene_token(genes)))
        records.append({
            "axis": axis,
            "submodule_id": sid,
            "display_label": label,
            "n_genes": len(gene_list),
            "genes": gene_list,
            "biological_readout": readout,
            "priority_weight": weight,
            "proposed_figure_family": figure,
        })
    return records


def classify_coverage(n_total: int, n_detected: int) -> str:
    if n_total <= 0:
        return "unresolved"
    frac = n_detected / n_total
    if n_total >= 10:
        if n_detected >= 8 and frac >= 0.65:
            return "adequate"
        if n_detected >= 5 and frac >= 0.40:
            return "partial"
    elif n_total >= 5:
        if n_detected >= 4 and frac >= 0.70:
            return "adequate"
        if n_detected >= 3 and frac >= 0.45:
            return "partial"
    else:
        if n_detected >= 3 and frac >= 0.75:
            return "adequate"
        if n_detected >= 2 and frac >= 0.50:
            return "partial"
    return "weak"


def write_tsv(path: Path, rows: List[Dict[str, object]], fieldnames: Sequence[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames, delimiter="\t", extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            clean = {}
            for key in fieldnames:
                value = row.get(key, "")
                if isinstance(value, list):
                    value = ";".join(str(x) for x in value)
                clean[key] = value
            writer.writerow(clean)


def write_gmt(path: Path, records: List[Dict[str, object]]) -> None:
    with path.open("w", encoding="utf-8") as handle:
        for record in records:
            desc = f"U26A|{record['axis']}|{record['biological_readout']}"
            handle.write("\t".join([str(record["submodule_id"]), desc] + list(record["genes"])) + "\n")


def robust_mean(values: List[float]) -> float:
    vals = [x for x in values if not math.isnan(x)]
    return statistics.mean(vals) if vals else float("nan")


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--project-root", default="__UTI_HOSTOMICS_PROJECT_ROOT__")
    parser.add_argument("--max-files-per-dataset", type=int, default=40)
    parser.add_argument("--output-tag", default="phaseU26A_expanded_endocrine_metabolic_immune_feasibility")
    args = parser.parse_args()

    root = Path(args.project_root).expanduser().resolve()
    if not root.exists():
        eprint(f"ERROR: project root does not exist: {root}")
        return 2

    metadata_dir = root / "03_metadata" / args.output_tag
    results_dir = root / "05_results" / args.output_tag
    tables_dir = root / "06_tables" / args.output_tag
    logs_dir = root / "08_logs" / args.output_tag
    for d in (metadata_dir, results_dir, tables_dir, logs_dir):
        d.mkdir(parents=True, exist_ok=True)

    records = submodule_records()
    library_rows = []
    for r in records:
        library_rows.append({
            "axis": r["axis"],
            "submodule_id": r["submodule_id"],
            "display_label": r["display_label"],
            "n_genes": r["n_genes"],
            "genes": ";".join(r["genes"]),
            "biological_readout": r["biological_readout"],
            "priority_weight": r["priority_weight"],
            "proposed_figure_family": r["proposed_figure_family"],
        })
    library_file = metadata_dir / "UTI_HostOmics_U26A_expanded_submodule_library.tsv"
    gmt_file = metadata_dir / "UTI_HostOmics_U26A_expanded_submodules.gmt"
    write_tsv(library_file, library_rows, [
        "axis", "submodule_id", "display_label", "n_genes", "genes",
        "biological_readout", "priority_weight", "proposed_figure_family",
    ])
    write_gmt(gmt_file, records)

    index_rows = []
    for item in INDEX_BLUEPRINTS:
        index_rows.append({
            "index_id": item["index_id"],
            "positive_modules": ";".join(item["positive_modules"]),
            "negative_modules": ";".join(item["negative_modules"]),
            "interpretation": item["interpretation"],
            "implementation_phase": "U26B or later after expression-matrix and contrast confirmation",
        })
    write_tsv(metadata_dir / "UTI_HostOmics_U26A_composite_index_blueprints.tsv", index_rows, [
        "index_id", "positive_modules", "negative_modules", "interpretation", "implementation_phase",
    ])

    r_helper = logs_dir / "U26A_extract_R_object_features_and_metadata.R"
    write_r_helper(r_helper)

    gtf_map = parse_gtf_mappings(root)
    dataset_universes: Dict[str, Set[str]] = {}
    dataset_meta: Dict[str, List[Dict[str, object]]] = defaultdict(list)
    source_inventory: List[Dict[str, object]] = []
    unresolved: List[Dict[str, object]] = []

    for dataset_id, spec in DATASETS.items():
        print(f"[U26A] Discovering files for {dataset_id} ({spec['label']})")
        candidates = discover_dataset_files(root, dataset_id, spec)
        universe: Set[str] = set()
        selected = candidates[: max(1, args.max_files_per_dataset)]
        if not selected:
            unresolved.append({
                "dataset": dataset_id,
                "issue": "No accession-matched or high-confidence alias-matched input files were discovered",
                "recommended_action": "locate the processed expression object or gene list and rerun U26A",
            })
        for path in selected:
            try:
                size_bytes = path.stat().st_size
            except OSError:
                size_bytes = 0
            # Skip extremely large raw matrices unless they are structured objects or feature files.
            if size_bytes > 8_000_000_000 and path.suffix.lower() not in {".h5ad", ".rds", ".rda", ".rdata"} and "feature" not in path.name.lower():
                source_inventory.append({
                    "dataset": dataset_id, "path": str(path), "size_bytes": size_bytes,
                    "candidate_score": candidate_score(path), "genes_extracted": 0,
                    "metadata_levels_extracted": 0, "status": "skipped_extremely_large_raw_file",
                })
                continue
            genes, meta, status = extract_file(path, r_helper)
            universe.update(genes)
            dataset_meta[dataset_id].extend(meta)
            # For likely metadata tables, inspect categorical columns separately.
            if any(key in path.name.lower() for key in META_KEYWORDS) and path.suffix.lower() in {".tsv", ".csv", ".txt", ".gz"}:
                dataset_meta[dataset_id].extend(read_metadata_table(path))
            source_inventory.append({
                "dataset": dataset_id,
                "path": str(path),
                "size_bytes": size_bytes,
                "candidate_score": candidate_score(path),
                "genes_extracted": len(genes),
                "metadata_levels_extracted": len(meta),
                "status": status,
            })
        dataset_universes[dataset_id] = universe

    all_ensembl = {g for genes in dataset_universes.values() for g in genes if is_ensembl(g)}
    orgdb_map = map_ensembl_with_orgdb(all_ensembl - set(gtf_map), logs_dir)
    ens_map = {**gtf_map, **orgdb_map}
    for dataset_id, genes in dataset_universes.items():
        mapped = {ens_map[g] for g in genes if g in ens_map}
        dataset_universes[dataset_id] = {g for g in genes if not is_ensembl(g)} | mapped

    write_tsv(tables_dir / "UTI_HostOmics_U26A_dataset_source_inventory.tsv", source_inventory, [
        "dataset", "path", "size_bytes", "candidate_score", "genes_extracted",
        "metadata_levels_extracted", "status",
    ])
    write_tsv(tables_dir / "UTI_HostOmics_U26A_unresolved_input_items.tsv", unresolved, [
        "dataset", "issue", "recommended_action",
    ])

    universe_rows = []
    for dataset_id, genes in dataset_universes.items():
        out_gene_file = metadata_dir / f"{dataset_id}_U26A_detected_gene_universe.txt"
        out_gene_file.write_text("\n".join(sorted(genes)) + ("\n" if genes else ""), encoding="utf-8")
        universe_rows.append({
            "dataset": dataset_id,
            "dataset_label": DATASETS[dataset_id]["label"],
            "layer": DATASETS[dataset_id]["layer"],
            "n_detected_gene_symbols": len(genes),
            "gene_universe_file": str(out_gene_file),
            "status": "resolved" if len(genes) >= 500 else ("limited" if genes else "unresolved"),
        })
    write_tsv(tables_dir / "UTI_HostOmics_U26A_dataset_gene_universe_summary.tsv", universe_rows, [
        "dataset", "dataset_label", "layer", "n_detected_gene_symbols", "gene_universe_file", "status",
    ])

    coverage_rows: List[Dict[str, object]] = []
    for record in records:
        genes = set(record["genes"])
        for dataset_id, universe in dataset_universes.items():
            detected = sorted(genes.intersection(universe))
            missing = sorted(genes.difference(universe))
            status = classify_coverage(len(genes), len(detected)) if universe else "unresolved"
            coverage_rows.append({
                "dataset": dataset_id,
                "dataset_label": DATASETS[dataset_id]["label"],
                "layer": DATASETS[dataset_id]["layer"],
                "axis": record["axis"],
                "submodule_id": record["submodule_id"],
                "display_label": record["display_label"],
                "n_module_genes": len(genes),
                "n_detected": len(detected),
                "coverage_fraction": round(len(detected) / len(genes), 4) if genes else "",
                "coverage_class": status,
                "detected_genes": ";".join(detected),
                "missing_genes": ";".join(missing),
                "proposed_figure_family": record["proposed_figure_family"],
            })
    write_tsv(tables_dir / "UTI_HostOmics_U26A_submodule_coverage_by_dataset.tsv", coverage_rows, [
        "dataset", "dataset_label", "layer", "axis", "submodule_id", "display_label",
        "n_module_genes", "n_detected", "coverage_fraction", "coverage_class",
        "detected_genes", "missing_genes", "proposed_figure_family",
    ])

    by_submodule: Dict[str, List[Dict[str, object]]] = defaultdict(list)
    for row in coverage_rows:
        by_submodule[str(row["submodule_id"])].append(row)
    submodule_recs: List[Dict[str, object]] = []
    for record in records:
        sid = str(record["submodule_id"])
        rows = by_submodule[sid]
        adequate_bulk = sum(1 for r in rows if r["layer"] == "bulk" and r["coverage_class"] == "adequate")
        partial_bulk = sum(1 for r in rows if r["layer"] == "bulk" and r["coverage_class"] == "partial")
        sc_status = next((str(r["coverage_class"]) for r in rows if r["layer"] == "single-cell"), "unresolved")
        numeric_cov = [float(r["coverage_fraction"]) for r in rows if r["coverage_fraction"] != ""]
        mean_cov = robust_mean(numeric_cov)
        if adequate_bulk >= 2 and sc_status == "adequate":
            recommendation = "main_figure_strong"
        elif adequate_bulk >= 2 and sc_status in {"partial", "unresolved"}:
            recommendation = "main_figure_candidate_with_bulk_emphasis"
        elif adequate_bulk + partial_bulk >= 2 and sc_status in {"adequate", "partial"}:
            recommendation = "supplement_or_secondary_main_panel"
        elif adequate_bulk + partial_bulk >= 1:
            recommendation = "supplement_exploratory"
        else:
            recommendation = "defer_or_require_additional_dataset"
        submodule_recs.append({
            "axis": record["axis"],
            "submodule_id": sid,
            "display_label": record["display_label"],
            "adequate_bulk_datasets": adequate_bulk,
            "partial_bulk_datasets": partial_bulk,
            "single_cell_coverage": sc_status,
            "mean_coverage_fraction": "" if math.isnan(mean_cov) else round(mean_cov, 4),
            "priority_weight": record["priority_weight"],
            "proposed_figure_family": record["proposed_figure_family"],
            "figure_recommendation": recommendation,
        })
    write_tsv(tables_dir / "UTI_HostOmics_U26A_submodule_figure_priority.tsv", submodule_recs, [
        "axis", "submodule_id", "display_label", "adequate_bulk_datasets", "partial_bulk_datasets",
        "single_cell_coverage", "mean_coverage_fraction", "priority_weight",
        "proposed_figure_family", "figure_recommendation",
    ])

    axis_rows: List[Dict[str, object]] = []
    by_axis: Dict[str, List[Dict[str, object]]] = defaultdict(list)
    for row in submodule_recs:
        by_axis[str(row["axis"])].append(row)
    axis_feasibility_score: Dict[str, float] = {}
    for axis, rows in by_axis.items():
        strong = sum(1 for r in rows if str(r["figure_recommendation"]).startswith("main_figure"))
        secondary = sum(1 for r in rows if r["figure_recommendation"] == "supplement_or_secondary_main_panel")
        exploratory = sum(1 for r in rows if r["figure_recommendation"] == "supplement_exploratory")
        deferred = sum(1 for r in rows if r["figure_recommendation"] == "defer_or_require_additional_dataset")
        score = (strong + 0.65 * secondary + 0.30 * exploratory) / max(1, len(rows))
        axis_feasibility_score[axis] = score
        if score >= 0.75:
            axis_rec = "main_figure_axis"
        elif score >= 0.50:
            axis_rec = "main_figure_with_selected_submodules"
        elif score >= 0.25:
            axis_rec = "supplementary_axis"
        else:
            axis_rec = "defer_or_expand_dataset_support"
        figures = sorted({str(r["proposed_figure_family"]) for r in rows})
        axis_rows.append({
            "axis": axis,
            "n_submodules": len(rows),
            "n_main_candidates": strong,
            "n_secondary_candidates": secondary,
            "n_exploratory": exploratory,
            "n_deferred": deferred,
            "feasibility_score": round(score, 4),
            "axis_recommendation": axis_rec,
            "proposed_figure_family": ";".join(figures),
        })
    axis_rows.sort(key=lambda x: (-float(x["feasibility_score"]), str(x["axis"])))
    write_tsv(tables_dir / "UTI_HostOmics_U26A_axis_feasibility_summary.tsv", axis_rows, [
        "axis", "n_submodules", "n_main_candidates", "n_secondary_candidates", "n_exploratory",
        "n_deferred", "feasibility_score", "axis_recommendation", "proposed_figure_family",
    ])

    contrast_rows: List[Dict[str, object]] = []
    for dataset_id, meta in dataset_meta.items():
        contrast_rows.extend(metadata_to_contrasts(dataset_id, meta, "auto-discovered metadata"))
    for dataset_id, spec in DATASETS.items():
        if not any(r["dataset"] == dataset_id for r in contrast_rows):
            for contrast in spec["expected_contrasts"]:
                contrast_rows.append({
                    "dataset": dataset_id,
                    "source": "project handoff expectation",
                    "metadata_column": "not_auto_resolved",
                    "levels": contrast,
                    "n_levels": "",
                    "total_n": "",
                    "minimum_level_n": "",
                    "support_class": "requires_metadata_confirmation",
                    "recommended_statistics": "effect sizes and FDR first; use logistic regression/odds ratios only for adequate true binary outcomes",
                })
    write_tsv(tables_dir / "UTI_HostOmics_U26A_feasible_contrast_map.tsv", contrast_rows, [
        "dataset", "source", "metadata_column", "levels", "n_levels", "total_n",
        "minimum_level_n", "support_class", "recommended_statistics",
    ])

    rq_rows: List[Dict[str, object]] = []
    for rq in RESEARCH_QUESTIONS:
        scores = [axis_feasibility_score.get(axis, 0.0) for axis in rq["axes"]]
        feasibility = robust_mean(scores)
        rank_score = 0.6 * (float(rq["impact"]) / 5.0) + 0.4 * (0.0 if math.isnan(feasibility) else feasibility)
        if feasibility >= 0.70:
            testability = "testable_now_high_priority"
        elif feasibility >= 0.45:
            testability = "testable_now_with_selected_submodules"
        elif feasibility >= 0.20:
            testability = "exploratory_or_supplementary"
        else:
            testability = "requires_additional_dataset_or_gene_resolution"
        rq_rows.append({
            "research_question_id": rq["id"],
            "research_question": rq["question"],
            "axes": ";".join(rq["axes"]),
            "scientific_impact_weight": rq["impact"],
            "coverage_feasibility_score": round(0.0 if math.isnan(feasibility) else feasibility, 4),
            "combined_priority_score": round(rank_score, 4),
            "testability_class": testability,
            "proposed_figure": rq["figure"],
        })
    rq_rows.sort(key=lambda x: (-float(x["combined_priority_score"]), str(x["research_question_id"])))
    for idx, row in enumerate(rq_rows, start=1):
        row["rank"] = idx
    write_tsv(tables_dir / "UTI_HostOmics_U26A_high_impact_research_question_ranking.tsv", rq_rows, [
        "rank", "research_question_id", "research_question", "axes", "scientific_impact_weight",
        "coverage_feasibility_score", "combined_priority_score", "testability_class", "proposed_figure",
    ])

    main_candidates = [r for r in submodule_recs if str(r["figure_recommendation"]).startswith("main_figure")]
    supplement_candidates = [r for r in submodule_recs if r["figure_recommendation"] in {"supplement_or_secondary_main_panel", "supplement_exploratory"}]
    deferred = [r for r in submodule_recs if r["figure_recommendation"] == "defer_or_require_additional_dataset"]

    report = results_dir / "UTI_HostOmics_U26A_feasibility_report.md"
    with report.open("w", encoding="utf-8") as handle:
        handle.write("# Phase U26A - Expanded endocrine-metabolic-immune submodule feasibility report\n\n")
        handle.write(f"- Script version: `{SCRIPT_VERSION}`\n")
        handle.write(f"- Project root: `{root}`\n")
        handle.write(f"- Submodules defined: **{len(records)}**\n")
        handle.write(f"- Biological axes represented: **{len(by_axis)}**\n")
        handle.write("- Manuscript and existing figures were not modified.\n\n")

        handle.write("## Dataset gene-universe resolution\n\n")
        for row in universe_rows:
            handle.write(f"- **{row['dataset']}** ({row['dataset_label']}; {row['layer']}): {row['n_detected_gene_symbols']} gene symbols; status `{row['status']}`.\n")
        handle.write("\n")

        handle.write("## Axis-level feasibility\n\n")
        for row in axis_rows:
            handle.write(
                f"- **{row['axis']}**: score {row['feasibility_score']}; recommendation `{row['axis_recommendation']}`; "
                f"main candidates {row['n_main_candidates']}/{row['n_submodules']}.\n"
            )
        handle.write("\n")

        handle.write("## Main-figure candidates\n\n")
        if main_candidates:
            for row in sorted(main_candidates, key=lambda x: (str(x["proposed_figure_family"]), str(x["axis"]), str(x["submodule_id"]))):
                handle.write(f"- `{row['submodule_id']}` - {row['display_label']} ({row['proposed_figure_family']}; {row['figure_recommendation']}).\n")
        else:
            handle.write("- No main-figure candidate could be confirmed because one or more dataset gene universes were unresolved.\n")
        handle.write("\n")

        handle.write("## Supplementary or secondary candidates\n\n")
        for row in sorted(supplement_candidates, key=lambda x: (str(x["axis"]), str(x["submodule_id"])))[:80]:
            handle.write(f"- `{row['submodule_id']}` - {row['display_label']} ({row['figure_recommendation']}).\n")
        if len(supplement_candidates) > 80:
            handle.write(f"- Additional candidates are listed in the full table ({len(supplement_candidates) - 80} more).\n")
        handle.write("\n")

        handle.write("## Deferred or additional-dataset candidates\n\n")
        if deferred:
            for row in sorted(deferred, key=lambda x: (str(x["axis"]), str(x["submodule_id"]))):
                handle.write(f"- `{row['submodule_id']}` - {row['display_label']}.\n")
        else:
            handle.write("- None based on gene-universe coverage.\n")
        handle.write("\n")

        handle.write("## Ranked high-impact research questions\n\n")
        for row in rq_rows:
            handle.write(
                f"{row['rank']}. **{row['research_question_id']}** ({row['testability_class']}; score {row['combined_priority_score']}): "
                f"{row['research_question']}\n"
            )
        handle.write("\n")

        handle.write("## Statistical decision rules for the next phase\n\n")
        handle.write("- Use differential submodule activity with effect sizes and FDR-controlled inference as the default.\n")
        handle.write("- Use logistic regression and odds ratios only when a true binary outcome is present and group sizes are adequate.\n")
        handle.write("- Prefer AUROC with permutation testing when odds ratios are unstable or sample sizes are modest.\n")
        handle.write("- Build module-module and partial-correlation networks by tissue/context rather than pooling biologically incompatible layers.\n")
        handle.write("- In single-cell analyses, validate pathway activity within cell populations and at sample-aggregated pseudobulk level.\n")
        handle.write("- Describe glycolysis and related transcriptomic outputs as transcriptionally inferred metabolic pathway activity, not actual flux.\n\n")

        handle.write("## Immediate U26B entry criterion\n\n")
        handle.write("Proceed to U26B expression-level scoring when each intended dataset has a resolved expression object, sample metadata, and a confirmed contrast map. The U26A tables identify unresolved inputs and the strongest figure-ready axes.\n")

    manifest = {
        "script_version": SCRIPT_VERSION,
        "project_root": str(root),
        "outputs": {
            "submodule_library": str(library_file),
            "gmt": str(gmt_file),
            "coverage": str(tables_dir / "UTI_HostOmics_U26A_submodule_coverage_by_dataset.tsv"),
            "contrasts": str(tables_dir / "UTI_HostOmics_U26A_feasible_contrast_map.tsv"),
            "axis_summary": str(tables_dir / "UTI_HostOmics_U26A_axis_feasibility_summary.tsv"),
            "figure_priority": str(tables_dir / "UTI_HostOmics_U26A_submodule_figure_priority.tsv"),
            "research_questions": str(tables_dir / "UTI_HostOmics_U26A_high_impact_research_question_ranking.tsv"),
            "report": str(report),
        },
        "dataset_gene_counts": {k: len(v) for k, v in dataset_universes.items()},
        "n_submodules": len(records),
        "n_axes": len(by_axis),
    }
    (results_dir / "UTI_HostOmics_U26A_run_manifest.json").write_text(json.dumps(manifest, indent=2) + "\n", encoding="utf-8")

    print("[U26A] Completed.")
    print(f"[U26A] Report: {report}")
    print(f"[U26A] Coverage table: {tables_dir / 'UTI_HostOmics_U26A_submodule_coverage_by_dataset.tsv'}")
    print(f"[U26A] Contrast map: {tables_dir / 'UTI_HostOmics_U26A_feasible_contrast_map.tsv'}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
